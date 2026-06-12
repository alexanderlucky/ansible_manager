import json
import io
import os
import re
import subprocess
import threading
import queue
import time
import uuid
import zipfile
from pathlib import Path
from datetime import datetime
from django.conf import settings
from django.http import JsonResponse, StreamingHttpResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.shortcuts import render

PLAYBOOKS_DIR = Path(settings.PLAYBOOKS_DIR)

# Global run queue and state
# Each item: {'name': str, 'flags': ['-v',...], 'extra_vars': {'key':'val',...}}
run_queue = []
run_queue_lock = threading.Lock()
current_run = {'active': False, 'playbook': None, 'output': [], 'status': 'idle', 'flags': [], 'extra_vars': {}}
output_subscribers = []
output_lock = threading.Lock()
worker_thread = None

# Stats store: {playbook_name: [run_record, ...]}
# run_record = {timestamp, hosts: {hostname: {ok, changed, unreachable, failed, skipped, rescued}}}
stats_store = {}
stats_lock = threading.Lock()

DEFAULT_YML = """\
---
- name: Install VLC Package
  hosts: all
  tasks:
    - name: Install VLC
      apt:
        name: vlc
        state: present
"""

DEFAULT_INI = """\
[alt_nodes]
node1 ansible_host=localhost ansible_port=2201
node2 ansible_host=localhost ansible_port=2202
node34 ansible_host=10.249.11.1 ansible_port=2203
node18 ansible_host=10.249.11.9 ansible_port=2201
node277 ansible_host=10.249.11.7 ansible_port=2202
node334 ansible_host=10.249.11.1 ansible_port=2203

[all:vars]
ansible_user=root
ansible_ssh_pass=test
ansible_connection=ssh
ansible_ssh_common_args='-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null'
"""

# Regex: "hostname : ok=2  changed=0  unreachable=0  failed=1  skipped=0  rescued=0  ignored=0"
RECAP_RE = re.compile(
    r'^(?P<host>\S+)\s*:\s*'
    r'ok=(?P<ok>\d+)\s+'
    r'changed=(?P<changed>\d+)\s+'
    r'unreachable=(?P<unreachable>\d+)\s+'
    r'failed=(?P<failed>\d+)'
    r'(?:\s+skipped=(?P<skipped>\d+))?'
    r'(?:\s+rescued=(?P<rescued>\d+))?'
)




# ── SCHEDULER ──────────────────────────────────────────────────────────────────
# schedules: list of dicts {id, playbook, flags, extra_vars, mode, run_at, cron, enabled, last_run, next_run_display}
# mode: 'once' | 'cron'
schedules_store = []
schedules_lock  = threading.Lock()
_sched_thread   = None

def _parse_cron_next(cron_str):
    """Return next datetime string for a 5-field cron expression using stdlib only."""
    import calendar
    fields = cron_str.strip().split()
    if len(fields) != 5:
        return None
    now = datetime.now().replace(second=0, microsecond=0)
    # Try up to 366 days ahead to find next match
    from datetime import timedelta
    candidate = now + timedelta(minutes=1)
    for _ in range(366 * 24 * 60):
        m_ok  = _cron_field(fields[0], candidate.minute, 0, 59)
        h_ok  = _cron_field(fields[1], candidate.hour, 0, 23)
        d_ok  = _cron_field(fields[2], candidate.day, 1, 31)
        mo_ok = _cron_field(fields[3], candidate.month, 1, 12)
        w_ok  = _cron_field(fields[4], candidate.weekday(), 0, 6)  # 0=Mon
        if m_ok and h_ok and d_ok and mo_ok and w_ok:
            return candidate.strftime('%Y-%m-%d %H:%M')
        candidate += timedelta(minutes=1)
    return None

def _cron_field(expr, value, lo, hi):
    if expr == '*': return True
    if '/' in expr:
        parts = expr.split('/')
        step = int(parts[1])
        start = lo if parts[0] == '*' else int(parts[0])
        return (value - start) % step == 0 and value >= start
    if ',' in expr:
        return value in [int(x) for x in expr.split(',')]
    if '-' in expr:
        a, b = expr.split('-')
        return int(a) <= value <= int(b)
    return value == int(expr)

def _scheduler_loop():
    global _sched_thread
    while True:
        time.sleep(10)
        now = datetime.now().replace(second=0, microsecond=0)
        with schedules_lock:
            jobs = list(schedules_store)
        for job in jobs:
            if not job.get('enabled'):
                continue
            fire = False
            if job['mode'] == 'once':
                try:
                    run_at = datetime.strptime(job['run_at'], '%Y-%m-%dT%H:%M')
                    if run_at.replace(second=0, microsecond=0) == now:
                        fire = True
                except Exception:
                    pass
            elif job['mode'] == 'cron':
                nxt = _parse_cron_next(job.get('cron', ''))
                if nxt and datetime.strptime(nxt, '%Y-%m-%d %H:%M') == now:
                    fire = True
            if fire:
                job['last_run'] = now.strftime('%Y-%m-%d %H:%M')
                if job['mode'] == 'once':
                    job['enabled'] = False  # auto-disable after firing
                item = {'name': job['playbook'], 'flags': job.get('flags', []), 'extra_vars': job.get('extra_vars', {})}
                with run_queue_lock:
                    run_queue.append(item)
                _ensure_worker()
                _broadcast(f"[QUEUE] Scheduled run fired: {job['playbook']}\n")

def _ensure_scheduler():
    global _sched_thread
    if _sched_thread is None or not _sched_thread.is_alive():
        _sched_thread = threading.Thread(target=_scheduler_loop, daemon=True)
        _sched_thread.start()

_ensure_scheduler()

# ── VIEWS ──────────────────────────────────────────────────────────────────────

def index(request):
    return render(request, 'playbooks_app/index.html')


def _build_tree(path, base):
    """Recursively build file/folder tree. relpath is relative to the playbook root (base)."""
    nodes = []
    try:
        entries = sorted(path.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
    except PermissionError:
        return nodes
    for entry in entries:
        rel = str(entry.relative_to(base))
        if entry.is_dir():
            nodes.append({
                'type': 'dir',
                'name': entry.name,
                'relpath': rel,
                'children': _build_tree(entry, base),
            })
        else:
            nodes.append({
                'type': 'file',
                'name': entry.name,
                'relpath': rel,
                'ext': entry.suffix.lstrip('.') or 'txt',
            })
    return nodes


def list_playbooks(request):
    playbooks = {}
    if PLAYBOOKS_DIR.exists():
        for item in sorted(PLAYBOOKS_DIR.iterdir()):
            if item.is_dir():
                playbooks[item.name] = _build_tree(item, item)
    return JsonResponse({'playbooks': playbooks})


@csrf_exempt
@require_http_methods(["POST"])
def create_playbook(request):
    data = json.loads(request.body)
    name = data.get('name', '').strip().replace(' ', '_')
    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    pb_dir = PLAYBOOKS_DIR / name
    if pb_dir.exists():
        return JsonResponse({'error': 'Playbook already exists'}, status=400)

    pb_dir.mkdir(parents=True)
    (pb_dir / 'playbook.yml').write_text(DEFAULT_YML)
    (pb_dir / 'inventory.ini').write_text(DEFAULT_INI)

    return JsonResponse({'success': True, 'name': name})


def _safe_path(playbook, relpath):
    """Resolve a path and ensure it stays inside the playbook dir. Returns Path or None."""
    pb_dir = PLAYBOOKS_DIR / playbook
    target = (pb_dir / relpath).resolve()
    try:
        target.relative_to(pb_dir.resolve())
        return target
    except ValueError:
        return None


@csrf_exempt
def get_file(request):
    playbook = request.GET.get('playbook')
    relpath  = request.GET.get('file')
    if not playbook or not relpath:
        return JsonResponse({'error': 'Missing params'}, status=400)

    filepath = _safe_path(playbook, relpath)
    if not filepath:
        return JsonResponse({'error': 'Invalid path'}, status=403)
    if not filepath.exists() or not filepath.is_file():
        return JsonResponse({'error': 'File not found'}, status=404)

    try:
        content = filepath.read_text(encoding='utf-8')
    except UnicodeDecodeError:
        return JsonResponse({'error': 'Binary file — cannot edit in browser'}, status=415)

    return JsonResponse({'content': content, 'filename': filepath.name})


@csrf_exempt
@require_http_methods(["POST"])
def save_file(request):
    data     = json.loads(request.body)
    playbook = data.get('playbook')
    relpath  = data.get('file')
    content  = data.get('content', '')

    if not playbook or not relpath:
        return JsonResponse({'error': 'Missing params'}, status=400)

    filepath = _safe_path(playbook, relpath)
    if not filepath:
        return JsonResponse({'error': 'Invalid path'}, status=403)

    filepath.parent.mkdir(parents=True, exist_ok=True)
    filepath.write_text(content, encoding='utf-8')
    return JsonResponse({'success': True})


# ── FILESYSTEM OPS ─────────────────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(["POST"])
def fs_create_file(request):
    data     = json.loads(request.body)
    playbook = data.get('playbook')
    relpath  = data.get('relpath', '').strip()
    if not playbook or not relpath:
        return JsonResponse({'error': 'Missing params'}, status=400)

    target = _safe_path(playbook, relpath)
    if not target:
        return JsonResponse({'error': 'Invalid path'}, status=403)
    if target.exists():
        return JsonResponse({'error': 'Already exists'}, status=400)

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text('', encoding='utf-8')
    return JsonResponse({'success': True})


@csrf_exempt
@require_http_methods(["POST"])
def fs_create_folder(request):
    data     = json.loads(request.body)
    playbook = data.get('playbook')
    relpath  = data.get('relpath', '').strip()
    if not playbook or not relpath:
        return JsonResponse({'error': 'Missing params'}, status=400)

    target = _safe_path(playbook, relpath)
    if not target:
        return JsonResponse({'error': 'Invalid path'}, status=403)
    if target.exists():
        return JsonResponse({'error': 'Already exists'}, status=400)

    target.mkdir(parents=True)
    return JsonResponse({'success': True})


@csrf_exempt
@require_http_methods(["POST"])
def fs_rename(request):
    import shutil
    data     = json.loads(request.body)
    playbook = data.get('playbook')
    relpath  = data.get('relpath', '').strip()
    newname  = data.get('newname', '').strip()
    if not playbook or not relpath or not newname:
        return JsonResponse({'error': 'Missing params'}, status=400)
    if '/' in newname or '\\' in newname:
        return JsonResponse({'error': 'New name must not contain path separators'}, status=400)

    src = _safe_path(playbook, relpath)
    if not src or not src.exists():
        return JsonResponse({'error': 'Source not found'}, status=404)

    dst = _safe_path(playbook, str(Path(relpath).parent / newname))
    if not dst:
        return JsonResponse({'error': 'Invalid destination'}, status=403)
    if dst.exists():
        return JsonResponse({'error': 'Name already taken'}, status=400)

    shutil.move(str(src), str(dst))
    return JsonResponse({'success': True, 'newrelpath': str(dst.relative_to(PLAYBOOKS_DIR / playbook))})


@csrf_exempt
@require_http_methods(["POST"])
def fs_delete(request):
    import shutil
    data     = json.loads(request.body)
    playbook = data.get('playbook')
    relpath  = data.get('relpath', '').strip()
    if not playbook or not relpath:
        return JsonResponse({'error': 'Missing params'}, status=400)

    target = _safe_path(playbook, relpath)
    if not target or not target.exists():
        return JsonResponse({'error': 'Not found'}, status=404)

    # Never allow deleting the playbook root itself
    pb_root = (PLAYBOOKS_DIR / playbook).resolve()
    if target.resolve() == pb_root:
        return JsonResponse({'error': 'Use delete playbook instead'}, status=400)

    if target.is_dir():
        shutil.rmtree(target)
    else:
        target.unlink()
    return JsonResponse({'success': True})


@csrf_exempt
@require_http_methods(["POST"])
def delete_playbook(request):
    data = json.loads(request.body)
    name = data.get('name')
    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    pb_dir = PLAYBOOKS_DIR / name
    if not pb_dir.exists():
        return JsonResponse({'error': 'Not found'}, status=404)

    import shutil
    shutil.rmtree(pb_dir)

    with run_queue_lock:
        run_queue[:] = [x for x in run_queue if x != name]

    return JsonResponse({'success': True})


def get_queue(request):
    with run_queue_lock:
        q = list(run_queue)
    return JsonResponse({
        'queue': q,
        'current': current_run.get('playbook'),
        'status': current_run.get('status', 'idle'),
        'current_flags': current_run.get('flags', []),
        'current_extra_vars': current_run.get('extra_vars', {}),
    })


@csrf_exempt
@require_http_methods(["POST"])
def add_to_queue(request):
    data = json.loads(request.body)
    name = data.get('name')
    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    pb_dir = PLAYBOOKS_DIR / name
    if not pb_dir.exists():
        return JsonResponse({'error': 'Playbook not found'}, status=404)

    # Validate flags — only allow known safe ones
    raw_flags = data.get('flags', [])
    allowed_flags = {'-v', '-vv', '-vvv', '-vvvv', '--check', '--diff', '--syntax-check', '--list-tasks', '--list-hosts', '--step'}
    flags = [f for f in raw_flags if f in allowed_flags]

    # extra_vars: dict of key->value strings
    extra_vars = {}
    raw_ev = data.get('extra_vars', {})
    if isinstance(raw_ev, dict):
        for k, v in raw_ev.items():
            k = str(k).strip()
            v = str(v).strip()
            if k and re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', k):
                extra_vars[k] = v

    item = {'name': name, 'flags': flags, 'extra_vars': extra_vars}

    with run_queue_lock:
        # Replace if already queued (allows updating flags)
        run_queue[:] = [i for i in run_queue if i['name'] != name]
        run_queue.append(item)

    _ensure_worker()
    return JsonResponse({'success': True, 'queue': [i['name'] for i in run_queue]})


@csrf_exempt
@require_http_methods(["POST"])
def remove_from_queue(request):
    data = json.loads(request.body)
    name = data.get('name')
    with run_queue_lock:
        run_queue[:] = [i for i in run_queue if i['name'] != name]
    return JsonResponse({'success': True, 'queue': [i['name'] for i in run_queue]})


# ── STATS ──────────────────────────────────────────────────────────────────────

def get_stats(request):
    playbook = request.GET.get('playbook')
    with stats_lock:
        if playbook:
            data = stats_store.get(playbook, [])
            return JsonResponse({'playbook': playbook, 'runs': data})
        # Return all
        return JsonResponse({'stats': dict(stats_store)})


def _record_stats(playbook_name, run_record):
    """Store a completed run's stats and patch the INI file."""
    with stats_lock:
        if playbook_name not in stats_store:
            stats_store[playbook_name] = []
        stats_store[playbook_name].append(run_record)

    # Collect failed/unreachable hosts
    bad_hosts = [
        h for h, v in run_record['hosts'].items()
        if int(v.get('failed', 0)) > 0 or int(v.get('unreachable', 0)) > 0
    ]

    if bad_hosts:
        _patch_ini_failed(playbook_name, bad_hosts)


def _patch_ini_failed(playbook_name, bad_hosts):
    """Append/replace [failed] group in the inventory INI file."""
    pb_dir = PLAYBOOKS_DIR / playbook_name
    ini_files = list(pb_dir.glob('*.ini')) + list(pb_dir.glob('*.cfg'))
    if not ini_files:
        return

    ini_path = ini_files[0]
    content = ini_path.read_text()

    # Remove existing [failed] section
    lines = content.splitlines()
    new_lines = []
    in_failed = False
    for line in lines:
        if line.strip() == '[failed]':
            in_failed = True
            continue
        if in_failed:
            # Stop skipping when we hit the next group header
            if line.strip().startswith('['):
                in_failed = False
            else:
                continue
        if not in_failed:
            new_lines.append(line)

    # Strip trailing blank lines then append new [failed] section
    while new_lines and not new_lines[-1].strip():
        new_lines.pop()

    new_lines.append('')
    new_lines.append('[failed]')
    for host in sorted(bad_hosts):
        new_lines.append(host)
    new_lines.append('')

    ini_path.write_text('\n'.join(new_lines))


# ── EXCEL EXPORT ───────────────────────────────────────────────────────────────

def export_stats_excel(request):
    playbook = request.GET.get('playbook')  # optional filter

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=500)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Color palette
    HDR_FILL  = PatternFill("solid", fgColor="0D1117")
    OK_FILL   = PatternFill("solid", fgColor="0D2818")
    FAIL_FILL = PatternFill("solid", fgColor="2D0A0A")
    UNREACH_FILL = PatternFill("solid", fgColor="2D1A00")
    ALT_FILL  = PatternFill("solid", fgColor="111820")
    WHITE     = Font(color="E2E8F0", bold=False)
    HDR_FONT  = Font(color="4ADE80", bold=True)
    FAIL_FONT = Font(color="F87171", bold=True)
    OK_FONT   = Font(color="4ADE80", bold=True)
    WARN_FONT = Font(color="F59E0B", bold=True)
    CYAN_FONT = Font(color="22D3EE", bold=True)
    THIN      = Side(style='thin', color='2A3040')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(cell, font=None):
        cell.fill = HDR_FILL
        cell.font = font or HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    def style_cell(cell, font=None, fill=None):
        cell.fill = fill or ALT_FILL
        cell.font = font or WHITE
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    with stats_lock:
        scope = {playbook: stats_store[playbook]} if playbook and playbook in stats_store else dict(stats_store)

    for pb_name, runs in scope.items():
        ws = wb.create_sheet(title=pb_name[:31])  # Excel sheet name limit
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "4ADE80"

        # ── Title row ──
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"Ansible Run Stats — {pb_name}"
        title_cell.font = Font(color="4ADE80", bold=True, size=13)
        title_cell.fill = PatternFill("solid", fgColor="0A0C0F")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28

        # ── Column headers ──
        headers = ['Run #', 'Timestamp', 'Host', 'OK', 'Changed', 'Unreachable', 'Failed', 'Skipped', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            style_header(cell)
        ws.row_dimensions[2].height = 20

        row = 3
        for run_idx, run in enumerate(runs, 1):
            ts = run.get('timestamp', '')
            hosts = run.get('hosts', {})
            if not hosts:
                # Empty run row
                cell = ws.cell(row=row, column=1, value=run_idx)
                style_cell(cell)
                cell = ws.cell(row=row, column=2, value=ts)
                style_cell(cell)
                ws.cell(row=row, column=3, value='(no hosts)').font = WHITE
                row += 1
                continue

            for h_idx, (host, vals) in enumerate(sorted(hosts.items())):
                ok          = int(vals.get('ok', 0))
                changed     = int(vals.get('changed', 0))
                unreachable = int(vals.get('unreachable', 0))
                failed      = int(vals.get('failed', 0))
                skipped     = int(vals.get('skipped', 0))

                if failed > 0:
                    status, sfont, sfill = 'FAILED', FAIL_FONT, FAIL_FILL
                elif unreachable > 0:
                    status, sfont, sfill = 'UNREACHABLE', WARN_FONT, UNREACH_FILL
                elif ok > 0 or changed > 0:
                    status, sfont, sfill = 'OK', OK_FONT, OK_FILL
                else:
                    status, sfont, sfill = 'SKIPPED', WHITE, ALT_FILL

                row_fill = sfill if (failed > 0 or unreachable > 0) else ALT_FILL

                data_row = [run_idx, ts, host, ok, changed, unreachable, failed, skipped, status]
                for col, val in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    if col == 9:
                        style_cell(cell, font=sfont, fill=sfill)
                    elif col in (6, 7) and val > 0:
                        style_cell(cell, font=FAIL_FONT if col == 7 else WARN_FONT, fill=row_fill)
                    elif col == 4 and val > 0:
                        style_cell(cell, font=OK_FONT, fill=row_fill)
                    else:
                        style_cell(cell, fill=row_fill)
                row += 1

        # ── Summary block ──
        row += 1
        ws.cell(row=row, column=1, value='SUMMARY').font = CYAN_FONT
        ws.cell(row=row, column=1).fill = HDR_FILL
        row += 1

        total_runs = len(runs)
        all_hosts = {}
        for run in runs:
            for host, vals in run.get('hosts', {}).items():
                if host not in all_hosts:
                    all_hosts[host] = {'ok': 0, 'changed': 0, 'unreachable': 0, 'failed': 0, 'runs': 0}
                all_hosts[host]['ok']          += int(vals.get('ok', 0))
                all_hosts[host]['changed']      += int(vals.get('changed', 0))
                all_hosts[host]['unreachable']  += int(vals.get('unreachable', 0))
                all_hosts[host]['failed']       += int(vals.get('failed', 0))
                all_hosts[host]['runs']         += 1

        sum_headers = ['Host', 'Total Runs', 'Total OK', 'Total Changed', 'Total Unreachable', 'Total Failed', 'Fail Rate']
        for col, h in enumerate(sum_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            style_header(cell)
        row += 1

        for host, agg in sorted(all_hosts.items()):
            fail_rate = f"{agg['failed']/agg['runs']*100:.0f}%" if agg['runs'] else '0%'
            summary_row = [host, agg['runs'], agg['ok'], agg['changed'], agg['unreachable'], agg['failed'], fail_rate]
            has_issues = agg['failed'] > 0 or agg['unreachable'] > 0
            for col, val in enumerate(summary_row, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col == 6 and val > 0:
                    style_cell(cell, font=FAIL_FONT, fill=FAIL_FILL)
                elif col == 5 and val > 0:
                    style_cell(cell, font=WARN_FONT, fill=UNREACH_FILL)
                elif col == 7 and has_issues:
                    style_cell(cell, font=FAIL_FONT, fill=FAIL_FILL)
                else:
                    style_cell(cell, fill=OK_FILL if not has_issues else ALT_FILL)
            row += 1

        # Column widths
        col_widths = [7, 22, 24, 8, 10, 14, 10, 10, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # If no data at all, add a placeholder sheet
    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws['A1'] = "No runs recorded yet."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ansible_stats_{playbook or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── RUN ENGINE ─────────────────────────────────────────────────────────────────

def _broadcast(line):
    with output_lock:
        current_run['output'].append(line)
        for q in output_subscribers:
            try:
                q.put_nowait(line)
            except Exception:
                pass


def _parse_recap(output_lines):
    """Extract per-host stats from PLAY RECAP section."""
    in_recap = False
    hosts = {}
    for line in output_lines:
        if 'PLAY RECAP' in line:
            in_recap = True
            continue
        if in_recap:
            m = RECAP_RE.match(line.strip())
            if m:
                hosts[m.group('host')] = {
                    'ok':          int(m.group('ok')),
                    'changed':     int(m.group('changed')),
                    'unreachable': int(m.group('unreachable')),
                    'failed':      int(m.group('failed')),
                    'skipped':     int(m.group('skipped') or 0),
                    'rescued':     int(m.group('rescued') or 0),
                }
    return hosts


def _run_playbook(name, flags=None, extra_vars=None):
    flags = flags or []
    extra_vars = extra_vars or {}

    pb_dir = PLAYBOOKS_DIR / name
    yml_files = list(pb_dir.glob('*.yml')) + list(pb_dir.glob('*.yaml'))
    ini_files = list(pb_dir.glob('*.ini')) + list(pb_dir.glob('*.cfg'))

    if not yml_files:
        _broadcast(f'[ERROR] No .yml file found in {name}\n')
        return

    playbook_file = yml_files[0]
    cmd = ['ansible-playbook', str(playbook_file)]

    if ini_files:
        cmd += ['-i', str(ini_files[0])]

    # Add verbosity / mode flags
    cmd += flags

    # Add extra vars
    if extra_vars:
        ev_str = ' '.join(f'{k}={v}' for k, v in extra_vars.items())
        cmd += ['-e', ev_str]

    _broadcast(f'[START] Running: {" ".join(cmd)}\n')

    collected_lines = []

    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            cwd=str(pb_dir),
        )
        for line in proc.stdout:
            collected_lines.append(line)
            _broadcast(line)
        proc.wait()
        _broadcast(f'[DONE] Exit code: {proc.returncode}\n')

    except FileNotFoundError:
        _broadcast('[WARN] ansible-playbook not found, running in demo mode\n')
        # Demo with mixed results for testing
        demo_lines = [
            f'PLAY [{name}] {"*"*40}\n',
            '\n',
            'TASK [Gathering Facts] ' + '*'*40 + '\n',
            'ok: [web1.example.com]\n',
            'ok: [web2.example.com]\n',
            'fatal: [db1.example.com]: UNREACHABLE!\n',
            '\n',
            'TASK [Print hello] ' + '*'*40 + '\n',
            'ok: [web1.example.com] => {"msg": "Hello from Ansible!"}\n',
            'changed: [web2.example.com] => {"msg": "Hello from Ansible!"}\n',
            '\n',
            'PLAY RECAP ' + '*'*40 + '\n',
            'web1.example.com          : ok=2    changed=0    unreachable=0    failed=0    skipped=0    rescued=0\n',
            'web2.example.com          : ok=1    changed=1    unreachable=0    failed=0    skipped=0    rescued=0\n',
            'db1.example.com           : ok=0    changed=0    unreachable=1    failed=0    skipped=0    rescued=0\n',
            '\n',
            '[DONE] Demo run complete.\n',
        ]
        for line in demo_lines:
            time.sleep(0.25)
            collected_lines.append(line)
            _broadcast(line)

    except Exception as e:
        _broadcast(f'[ERROR] {e}\n')

    # Parse recap and store stats
    hosts = _parse_recap(collected_lines)
    if hosts:
        run_record = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'hosts': hosts,
        }
        _record_stats(name, run_record)
        # Broadcast a special stats event so the frontend can update
        stats_summary = {h: v for h, v in hosts.items()}
        _broadcast(f'[STATS] {json.dumps(stats_summary)}\n')


def _worker():
    global worker_thread
    while True:
        with run_queue_lock:
            if not run_queue:
                current_run['active'] = False
                current_run['playbook'] = None
                current_run['status'] = 'idle'
                worker_thread = None
                return
            item = run_queue.pop(0)
        name       = item['name']
        flags      = item.get('flags', [])
        extra_vars = item.get('extra_vars', {})

        current_run['active']     = True
        current_run['playbook']   = name
        current_run['status']     = 'running'
        current_run['output']     = []
        current_run['flags']      = flags
        current_run['extra_vars'] = extra_vars
        _broadcast(f'[QUEUE] Starting: {name}\n')

        _run_playbook(name, flags=flags, extra_vars=extra_vars)

        current_run['status'] = 'idle'
        time.sleep(0.5)


def _ensure_worker():
    global worker_thread
    if worker_thread is None or not worker_thread.is_alive():
        worker_thread = threading.Thread(target=_worker, daemon=True)
        worker_thread.start()



# ── SCHEDULER VIEWS ───────────────────────────────────────────────────────────

def list_schedules(request):
    with schedules_lock:
        jobs = list(schedules_store)
    # Annotate next_run for cron jobs
    result = []
    for j in jobs:
        item = dict(j)
        if j['mode'] == 'cron':
            item['next_run_display'] = _parse_cron_next(j.get('cron', '')) or '?'
        elif j['mode'] == 'once':
            item['next_run_display'] = j.get('run_at', '').replace('T', ' ')
        result.append(item)
    return JsonResponse({'schedules': result})


@csrf_exempt
@require_http_methods(["POST"])
def add_schedule(request):
    data       = json.loads(request.body)
    playbook   = data.get('playbook', '').strip()
    mode       = data.get('mode', 'once')        # 'once' | 'cron'
    run_at     = data.get('run_at', '')           # ISO datetime for 'once'
    cron       = data.get('cron', '')             # 5-field cron for 'cron'
    raw_flags  = data.get('flags', [])
    raw_ev     = data.get('extra_vars', {})

    if not playbook:
        return JsonResponse({'error': 'playbook required'}, status=400)
    if not (PLAYBOOKS_DIR / playbook).exists():
        return JsonResponse({'error': 'playbook not found'}, status=404)
    if mode == 'once' and not run_at:
        return JsonResponse({'error': 'run_at required for once mode'}, status=400)
    if mode == 'cron':
        if not cron or len(cron.split()) != 5:
            return JsonResponse({'error': 'cron must be a 5-field expression'}, status=400)

    allowed_flags = {'-v', '-vv', '-vvv', '-vvvv', '--check', '--diff', '--syntax-check', '--list-tasks', '--list-hosts', '--step'}
    flags = [f for f in raw_flags if f in allowed_flags]
    extra_vars = {}
    if isinstance(raw_ev, dict):
        for k, v in raw_ev.items():
            k = str(k).strip()
            if k and re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', k):
                extra_vars[k] = str(v).strip()

    job = {
        'id':         str(uuid.uuid4())[:8],
        'playbook':   playbook,
        'mode':       mode,
        'run_at':     run_at,
        'cron':       cron,
        'flags':      flags,
        'extra_vars': extra_vars,
        'enabled':    True,
        'last_run':   None,
    }
    with schedules_lock:
        schedules_store.append(job)

    return JsonResponse({'success': True, 'id': job['id']})


@csrf_exempt
@require_http_methods(["POST"])
def remove_schedule(request):
    data = json.loads(request.body)
    jid  = data.get('id', '')
    with schedules_lock:
        before = len(schedules_store)
        schedules_store[:] = [j for j in schedules_store if j['id'] != jid]
        removed = len(schedules_store) < before
    return JsonResponse({'success': removed})


def stream_run(request):
    """SSE endpoint: streams live output to browser."""
    client_queue = queue.Queue()

    with output_lock:
        for line in current_run.get('output', []):
            client_queue.put(line)
        output_subscribers.append(client_queue)

    def event_stream():
        try:
            while True:
                try:
                    line = client_queue.get(timeout=20)
                    data = line.replace('\n', '\\n')
                    yield f'data: {data}\n\n'
                except queue.Empty:
                    yield ': keepalive\n\n'
        except GeneratorExit:
            pass
        finally:
            with output_lock:
                if client_queue in output_subscribers:
                    output_subscribers.remove(client_queue)

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

import json
import requests

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def chat(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    data = json.loads(request.body)

    response = requests.post(
        "http://192.168.18.199:11434/api/chat",
        json=data,
        timeout=300,
    )

    return JsonResponse(response.json(), safe=False)